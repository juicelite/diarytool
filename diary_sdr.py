#coding=utf-8
import sys
import configparser
import openpyxl
import codecs
import smtplib
import getpass
from datetime import date, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

diary_date = datetime.today()
diary_xl = ""
diary_tp = ""

mail_smtp = ""
mail_from = ""
mail_to = ""
mail_cc = ""
mail_subject = ""
mail_body = ""

def load_config(p_date):
	config = configparser.ConfigParser()
	config.read(["diary.properties"], "utf-8")
	global diary_xl, mail_smtp, mail_from, mail_to, mail_cc, mail_subject
	diary_xl = config["DailyReport"]["diaryxl"]
	mail_smtp = config["DailyReport"]["smtp"]
	mail_from = config["DailyReport"]["from"]
	mail_to = config["DailyReport"]["to"]
	mail_cc = config["DailyReport"]["cc"]
	mail_subject = config["DailyReport"]["subject"] + p_date.strftime("%Y-%m-%d")
	
def load_diary_template():
	global diary_tp
	with codecs.open("diary.template", "r", "utf-8") as f:
		diary_tp = f.read()
	
def conpose_mail(p_date):
	global mail_body
	delta = (p_date.date() - date(1900, 1, 1)).days + 2
	wb = openpyxl.load_workbook(diary_xl)
	ws = wb.worksheets[0]
	for i in xrange(2, ws.get_highest_row()):
		cell = ws.cell(row = i, column = 1)
		if(delta == cell.value):
			task_am = format_task(ws.cell(row = i, column = 3).value)
			task_pm = format_task(ws.cell(row = i, column = 5).value)
			task_ev = format_task(ws.cell(row = i, column = 7).value)
			task_ex = format_task(ws.cell(row = i, column = 9).value)
			date_str = u"%d月%d日" % (p_date.month, p_date.day)
			mail_body = diary_tp % (date_str, task_am, task_pm, task_ev, task_ex)
			break

def format_task(p_task):
	m_task = u""
	if(p_task != None):
		m_task = unicode(p_task).replace("\n", "<br>")
	return m_task
	
def send_mail(p_from, p_to, p_cc, p_subject, p_body):
	msg = MIMEMultipart("alternative")
	msg.set_charset("utf-8")
	msg["Subject"] = p_subject
	msg["From"] = p_from
	msg["To"] = p_to
	msg["Cc"] = p_cc
	msg.attach(MIMEText(p_body, "html", "utf-8"))
	
	output_mail_info(p_from, p_to, p_cc, p_subject, p_body)
	
	m_opt = ""
	while(m_opt != "n" and m_opt != "y" and m_opt != "l"):
		m_opt = raw_input("Send diary ? y:yes, n:no, l:login\n").lower()
	if(m_opt == "n"):
		print "User canceled sending diary!"
		return
	svr = smtplib.SMTP("corp.netease.com")
	if(m_opt == "l"):
		m_user = raw_input("Username: ")
		m_pswd = getpass.getpass()
		try:
			svr = smtplib.SMTP(mail_smtp)
			svr.login(m_user, m_pswd)
		except smtplib.SMTPAuthenticationError, err:
			print "Authentication failed!"
			return 1
		else:
			print "Authentication successfully!"
	try:
		svr.sendmail(p_from, [p_to] + p_cc.split(","), msg.as_string())
		svr.quit()
	except smtplib.SMTPRecipientsRefused, err:
		# sys.stderr.write("ERROR: %s\n" % str(err))
		return 1
	else:
		print "Diary was sent successfully!"
	

def output_mail_info(p_from, p_to, p_cc, p_subject, p_body):
	print "From: %s" % p_from
	print "Subject: %s" % p_subject
	print "To: %s" % p_to
	print "Cc: %s" % p_cc
	print "-------------------------------------------------------------------------------------------------------------------------------------------------------\n"
	print "%s\n\n" % p_body
	

def main():
	global diary_date
	if(len(sys.argv) >= 2):
		diary_date = datetime.strptime(sys.argv[1], "%Y-%m-%d")
	load_config(diary_date)
	load_diary_template()
	conpose_mail(diary_date)
	send_mail(mail_from, mail_to, mail_cc, mail_subject, mail_body)

if(__name__ == "__main__"):
	sys.exit(main())