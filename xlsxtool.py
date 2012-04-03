#coding=utf-8

import openpyxl
import collections
from datetime import date, datetime, timedelta

class DiaryXlsxUtil:
	
	diarydata = collections.namedtuple("DiaryData", ["date", "breakfast", "lunch", "dinner", "dessert"])
	
	BREAKFAST_ROW = 3
	LUNCH_ROW = 5
	DINNER_ROW = 7
	DESSERT_ROW = 9
	
	def __init__(self, diary_path):
		self.diary_path = diary_path
		self.diary_wb = openpyxl.load_workbook(diary_path)
		self.diary_ws = self.diary_wb.worksheets[0]
		
	def get_daily_data(self, p_date):
		r = self.get_diary_row(p_date)
		data = self._extract_diary(r, u"%d月%d日" % (p_date.month, p_date.day))
		return data
	
	def get_weekly_data(self, start_date, days):
		r = self.get_diary_row(start_date)
		diaries = list()
		m_delta = timedelta(days = 1)
		m_date = start_date.date()
		for i in xrange(days):
			data = self._extract_diary(r + i, u"%d月%d日" % (m_date.month, m_date.day))
			if(data):
				diaries.append(data)
			m_date = m_date + m_delta
		return diaries
	
	def overwrite(self, p_data, r, c):
		cell = self.diary_ws.cell(row = r, column = c)
		cell.value = p_data
		self._save_diary()
		
	def append(self, p_data, r, c):
		cell = self.diary_ws.cell(row = r, column = c)
		if(not cell.value):
			cell.value = p_data
		else:
			cell.value = u"%s\n%s" % (cell.value, p_data)
		self._save_diary()
		
	def get_diary_row(self, p_date):
		ws = self.diary_ws
		delta = self._days_from1900(p_date)
		for i in xrange(2, ws.get_highest_row()):
			cell = ws.cell(row = i, column = 1)
			if(delta == cell.value):
				return i
		else:
			return -1
			
	def _save_diary(self):
		self.diary_wb.save(self.diary_path)
			
	def _format_task(self, p_task):
		m_task = u""
		if(p_task != None):
			m_task = unicode(p_task).replace("\n", "<br>")
		return m_task
					
	def _days_from1900(self, p_date):
		return (p_date.date() - date(1900, 1, 1)).days + 2
	
	def _extract_diary(self, r, date_str):
		ws = self.diary_ws
		if(r != -1):
			task_am = self._format_task(ws.cell(row = r, column = DiaryXlsxUtil.BREAKFAST_ROW).value)
			task_pm = self._format_task(ws.cell(row = r, column = DiaryXlsxUtil.LUNCH_ROW).value)
			task_ev = self._format_task(ws.cell(row = r, column = DiaryXlsxUtil.DINNER_ROW).value)
			task_ex = self._format_task(ws.cell(row = r, column = DiaryXlsxUtil.DESSERT_ROW).value)
			return DiaryXlsxUtil.diarydata(date = date_str, breakfast = task_am, lunch = task_pm, dinner = task_ev, dessert = task_ex)
		else:
			return None